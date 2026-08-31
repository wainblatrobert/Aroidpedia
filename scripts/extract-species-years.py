# -*- coding: utf-8 -*-
"""
AROIDPEDIA - EXTRACT SPECIES YEARS   (v1, 2026-08-31)
Path in repo: scripts/extract-species-years.py

Reads the POWO genus exports off Google Drive and writes ONE committed file:

    data/species-years.json      {genus slug: {name, years: {year: [species]}}}

That file is the only thing the timeline builder needs, and it is the only
reason this script exists: the exports live on G:\\ and GitHub Actions cannot
see them, so the extraction happens here and the result is committed.

RUN IT when the POWO exports are refreshed:

    python scripts/extract-species-years.py

Then commit data/species-years.json. The Action rebuilds every timeline.

------------------------------------------------------------------
POWO IS THE BASE LIST, NOT THE VERDICT
------------------------------------------------------------------
House rule (2026-08-31): a published journal article trumps POWO. Where
Aroidpedia has ruled otherwise, the ruling lives in data/taxon-rulings.json
and is applied here — names POWO accepts can be excluded, names POWO has
dropped can be kept.

⚠ A ruling that no longer bites is a FAILURE, not a no-op. If POWO catches
up (starts listing an `include`, or drops an `exclude`) this script exits
non-zero and names the entry, so a stale ruling cannot sit in the file
pretending to do something. Delete it, or convert it, deliberately.

------------------------------------------------------------------
WHAT IT FILTERS, AND WHY
------------------------------------------------------------------
* Rows whose SPECIES NAME contains "x" (the multiplication sign). The export
  truncates hybrid rows to a bare "Alocasia \u00d7" / "Arum \u00d7" - 22 of them across
  the 153 files. Left in, the Alocasia timeline gains a milestone reading
  "Alocasia \u00d7" in 1965, which is what the live spec shows today.
* Rows with no year, or a year that will not parse to four digits.
* "~$" lock files and the two Araceae_Master_* workbooks.

\u26a0 SPECIES ONLY. Hybrids live on a separate sheet in these workbooks and are
deliberately not read: the timeline is the year each species was first
described, and a hybrid has no such date in the same sense.

\u26a0 THE EXPORTS LAG. Every genus file was written 2026-04; the master list
2026-07. Re-export before a big pass if the dates matter. This script reports
the oldest and newest file it read so that is never a surprise.
"""
import io, json, os, re, sys, datetime

sys.stdout.reconfigure(encoding="utf-8")

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("openpyxl is required:  pip install openpyxl")

EXCELS = os.environ.get(
    "AROIDPEDIA_EXCELS",
    r"G:\My Drive\PlantsV2\Aroidpedia\GENERA SPECIES LISTS\POWO_AUTOMATE"
    r"\Araceae_Exports\EXCELS",
)
REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(REPO, "data", "species-years.json")
RULINGS = os.path.join(REPO, "data", "taxon-rulings.json")

HYBRID = "\u00d7"
YEAR = re.compile(r"(1[5-9]\d\d|20\d\d)")


def read_genus(path):
    """-> (display name, {year: [species]}, notes) or None."""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if "SPECIES" not in wb.sheetnames:
            return None
        rows = list(wb["SPECIES"].iter_rows(values_only=True))
        if not rows:
            return None
        hdr = [str(h).strip() if h else "" for h in rows[0]]
        if "SPECIES NAME" not in hdr or "YEAR DESCRIBED" not in hdr:
            return None
        iN, iY = hdr.index("SPECIES NAME"), hdr.index("YEAR DESCRIBED")

        years, dropped_hybrid, dropped_year = {}, 0, 0
        for r in rows[1:]:
            name = str(r[iN]).strip() if r[iN] else ""
            if not name:
                continue
            if HYBRID in name:
                dropped_hybrid += 1
                continue
            m = YEAR.search(str(r[iY])) if r[iY] else None
            if not m:
                dropped_year += 1
                continue
            years.setdefault(m.group(1), []).append(name)
        for y in years:
            years[y] = sorted(set(years[y]))
        return years, dropped_hybrid, dropped_year
    finally:
        wb.close()


def apply_rulings(slug, years, rulings, log, problems):
    """Editorial rulings beat the export. Returns the amended year map."""
    r = rulings.get(slug)
    if not r:
        return years
    flat = {n: y for y, names in years.items() for n in names}

    for e in r.get("exclude", []):
        name = e["name"]
        if name not in flat:
            problems.append(
                "%s: exclude '%s' matches nothing — POWO already dropped it, "
                "so the ruling is stale. Remove it from taxon-rulings.json."
                % (slug, name))
            continue
        y = flat[name]
        years[y] = [n for n in years[y] if n != name]
        if not years[y]:
            del years[y]
        log.append("%s: excluded %s (%s)" % (slug, name, e.get("reason", "")[:60]))

    for e in r.get("include", []):
        name, y = e["name"], str(e["year"])
        if name in flat:
            problems.append(
                "%s: include '%s' is already in the export — POWO has caught "
                "up, so the ruling is stale. Remove it from taxon-rulings.json."
                % (slug, name))
            continue
        years.setdefault(y, [])
        if name not in years[y]:
            years[y] = sorted(years[y] + [name])
        log.append("%s: kept %s (%s), absent from POWO" % (slug, name, y))

    return years


def main():
    if not os.path.isdir(EXCELS):
        sys.exit("Exports not found: %s\nSet AROIDPEDIA_EXCELS to override." % EXCELS)

    files = sorted(
        f for f in os.listdir(EXCELS)
        if f.endswith(".xlsx") and not f.startswith("~$")
        and not f.startswith("Araceae_Master")
    )
    rulings = {}
    if os.path.exists(RULINGS):
        rulings = {k: v for k, v in
                   json.load(io.open(RULINGS, encoding="utf-8")).items()
                   if not k.startswith("_")}

    genera, skipped, stamps = {}, [], []
    ruling_log, ruling_problems = [], []
    tot_sp = tot_hy = tot_ny = 0

    for f in files:
        path = os.path.join(EXCELS, f)
        name = os.path.splitext(f)[0]
        try:
            got = read_genus(path)
        except Exception as e:
            skipped.append("%s :: %s" % (f, str(e)[:70]))
            continue
        if not got:
            skipped.append("%s :: no usable SPECIES sheet" % f)
            continue
        years, hy, ny = got
        years = apply_rulings(name.lower(), years, rulings,
                              ruling_log, ruling_problems)
        if not years:
            skipped.append("%s :: no dated species" % f)
            continue
        stamps.append(os.path.getmtime(path))
        tot_hy += hy
        tot_ny += ny
        tot_sp += sum(len(v) for v in years.values())
        genera[name.lower()] = {"name": name, "years": years}

    oldest = datetime.date.fromtimestamp(min(stamps)).isoformat()
    newest = datetime.date.fromtimestamp(max(stamps)).isoformat()

    doc = {
        "version": "1.0.0",
        "generated": datetime.date.today().isoformat(),
        "source": "POWO genus exports (Araceae_Exports/EXCELS), sheet SPECIES",
        "exportsOldest": oldest,
        "exportsNewest": newest,
        "note": "Species only. Hybrid rows and undated rows are excluded; see "
                "scripts/extract-species-years.py for the rules. Editorial "
                "rulings in data/taxon-rulings.json override the export.",
        "rulingsApplied": ruling_log,
        "generaCount": len(genera),
        "speciesCount": tot_sp,
        "genera": dict(sorted(genera.items())),
    }
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    io.open(OUT, "w", encoding="utf-8", newline="\n").write(
        json.dumps(doc, indent=1, ensure_ascii=False) + "\n")

    print("genera written    : %d" % len(genera))
    print("dated species     : %d" % tot_sp)
    print("hybrid rows cut   : %d" % tot_hy)
    print("undated rows cut  : %d" % tot_ny)
    print("export file dates : %s .. %s" % (oldest, newest))
    if newest < (datetime.date.today() - datetime.timedelta(days=120)).isoformat():
        print("!! the newest export is over 120 days old - consider re-exporting")
    if ruling_log:
        print("editorial rulings applied (%d):" % len(ruling_log))
        for r in ruling_log:
            print("  . " + r)
    if skipped:
        print("skipped:")
        for s in skipped:
            print("  ! " + s)
    print("->", OUT)

    if ruling_problems:
        print("\nSTALE RULINGS — the export has moved and these no longer bite:")
        for p in ruling_problems:
            print("  !! " + p)
        sys.exit(1)


if __name__ == "__main__":
    main()
