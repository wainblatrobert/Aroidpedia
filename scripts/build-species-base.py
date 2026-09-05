# -*- coding: utf-8 -*-
"""
AROIDPEDIA - BUILD SPECIES BASE   (v1, 2026-09-05)
Path in repo: scripts/build-species-base.py

    python scripts/build-species-base.py Cyrtosperma

Builds the STARTING species list for a genus from the two nomenclatural
sources the house treats as the base list - POWO and IPNI - and writes it
as a workbook in the same shape as the POWO genus exports on the Drive
(Araceae_Exports/EXCELS/<Genus>.xlsx), so it can sit in that folder and be
read by scripts/extract-species-years.py unchanged:

    data/species-base/<Genus>.xlsx            SPECIES / HYBRIDS / SYNONYMS /
                                              ALL NAMES / SOURCES
    data/species-base/<Genus>-species.csv     the SPECIES sheet, flat
    data/species-base/<Genus>-synonyms.csv    the SYNONYMS sheet, flat
    data/species-base/<Genus>-names.json      every name, reconciled, for code

------------------------------------------------------------------
WHERE THE DATA COMES FROM
------------------------------------------------------------------
IPNI   https://www.ipni.org/api/1/search?q=genus:<Genus>
       Every name ever published in the genus: authority, protologue
       (publication, collation, year), basionym, IPNI id, WFO id, BHL link,
       the original remarks (usually the type locality), and IPNI's own
       "inPowo" flag.

POWO   The World Checklist of Vascular Plants (WCVP) - the dataset POWO is
       built on - read through GBIF's copy of it:
           dataset f382f0ce-323a-4091-bb9f-add557f3a9a2, DOI 10.15468/6h8ucr
       It carries POWO's taxonomic status (Accepted / Synonym / Unplaced),
       the accepted name each synonym sinks into, the basionym, POWO's own
       range summary, the native/introduced TDWG level-3 distribution, and
       the POWO taxon URL (which embeds the IPNI id - that is the join key
       between the two sources).

       WHY THE MIRROR: powo.science.kew.org/api sits behind a Cloudflare
       browser challenge and answers 403 to scripts. The GBIF copy is the
       same WCVP release, refreshed when Kew publishes one; the SOURCES
       sheet records the release date so the lag is never a surprise. Every
       row carries its POWO URL so a name can be checked on POWO itself.

------------------------------------------------------------------
POWO IS THE BASE LIST, NOT THE VERDICT
------------------------------------------------------------------
Same house rule as the timelines (2026-08-31): a published journal article
trumps POWO. This script does NOT apply data/taxon-rulings.json - it is a
faithful pull of the two sources, so the editorial pass starts from what
the sources actually say. Rulings belong in the SPECIES sheet's editorial
columns (and in taxon-rulings.json once made).

------------------------------------------------------------------
WHAT GOES WHERE
------------------------------------------------------------------
SPECIES    one row per POWO-ACCEPTED species. Header names "SPECIES NAME"
           and "YEAR DESCRIBED" are exactly what extract-species-years.py
           looks for. Hybrids are NOT on this sheet (same rule as the
           exports: they live on HYBRIDS).
HYBRIDS    hybrid names (IPNI hybrid flag or a x in the name). Headers only
           when the genus has none.
SYNONYMS   every other name that touches the genus: genus names POWO sinks
           into an accepted species here or elsewhere, names POWO leaves
           unplaced, names only IPNI knows, and names from OTHER genera that
           POWO sinks into an accepted species of this genus.
ALL NAMES  the raw union of both sources, one row per name, with the join
           result - the audit trail for the sheets above.
SOURCES    endpoints, release dates, counts, and every discrepancy the join
           found (spelling and authority differences between IPNI and
           POWO, names in one source only).

Cache: raw API responses are cached under $AP_SPECIES_BASE_CACHE (default:
a folder in the system temp dir). Never commit the cache. Pass --refresh to
ignore it.

Requires: requests openpyxl
"""
import argparse, csv, datetime, hashlib, io, json, os, re, sys, tempfile, time

sys.stdout.reconfigure(encoding="utf-8")

try:
    import requests
except ImportError:
    sys.exit("requests is required:  pip install requests")
try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl is required:  pip install openpyxl")

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT_DIR = os.path.join(REPO, "data", "species-base")
CACHE = os.environ.get(
    "AP_SPECIES_BASE_CACHE",
    os.path.join(tempfile.gettempdir(), "aroidpedia-species-base-cache"))

IPNI_SEARCH = "https://www.ipni.org/api/1/search"
GBIF = "https://api.gbif.org/v1"
WCVP_DATASET = "f382f0ce-323a-4091-bb9f-add557f3a9a2"
WCVP_DOI = "10.15468/6h8ucr"
UA = "Aroidpedia species-base builder (github.com/wainblatrobert/Aroidpedia)"

HYBRID = "×"
YEAR_PAREN = re.compile(r"\((1[5-9]\d\d|20\d\d)\)")
YEAR_ANY = re.compile(r"\b(1[5-9]\d\d|20\d\d)\b")
IPNI_LSID = re.compile(r"names:(\d+-\d+)")

# --------------------------------------------------------------------------
# HTTP with a disk cache
# --------------------------------------------------------------------------
_session = requests.Session()
_session.headers["User-Agent"] = UA
_refresh = False
_calls = 0


def get_json(url, params=None):
    """GET url as JSON, cached on disk by the full URL."""
    global _calls
    # prepare_request, not Request.prepare(): only the former merges the
    # session headers, and IPNI's firewall rejects the default python UA.
    req = _session.prepare_request(requests.Request("GET", url, params=params))
    key = hashlib.sha1(req.url.encode("utf-8")).hexdigest()
    path = os.path.join(CACHE, key + ".json")
    if not _refresh and os.path.exists(path):
        with io.open(path, encoding="utf-8") as f:
            return json.load(f)
    for attempt in range(5):
        try:
            r = _session.send(req, timeout=60)
            if r.status_code == 200:
                data = r.json()
                os.makedirs(CACHE, exist_ok=True)
                with io.open(path, "w", encoding="utf-8") as f:
                    json.dump(data, f)
                _calls += 1
                time.sleep(0.15)
                return data
            if r.status_code in (429, 500, 502, 503, 504):
                time.sleep(2 ** attempt)
                continue
            r.raise_for_status()
        except requests.RequestException as e:
            if attempt == 4:
                raise
            print("  retry %d for %s (%s)" % (attempt + 1, req.url, e))
            time.sleep(2 ** attempt)
    raise RuntimeError("gave up on " + req.url)


# --------------------------------------------------------------------------
# IPNI
# --------------------------------------------------------------------------
def ipni_names(genus):
    out, page = [], 1
    while True:
        d = get_json(IPNI_SEARCH, {"q": "genus:" + genus, "perPage": 500, "page": page})
        rs = d.get("results", [])
        out.extend(r for r in rs if (r.get("genus") or "") == genus)
        if page >= int(d.get("totalPages") or 1) or not rs:
            break
        page += 1
    # IPNI returns one record per name id; keep the first if any repeats.
    seen, uniq = set(), []
    for r in out:
        if r["id"] in seen:
            continue
        seen.add(r["id"])
        uniq.append(r)
    return uniq


def ipni_protologue(r):
    """'Oesterr. Bot. Wochenbl. 7: 61 (1857)' from an IPNI record."""
    pub = (r.get("publication") or "").strip()
    col = (r.get("referenceCollation") or "").strip()
    yr = r.get("publicationYear")
    if pub:
        s = pub
        if col:
            s += " " + col
        if yr:
            s += " (%s)" % yr
        return s
    ref = (r.get("reference") or "").strip().rstrip(".")
    return ref


def ipni_basionym(r):
    b = (r.get("basionymStr") or "").strip()
    fam = (r.get("family") or "").strip()
    if fam and b.startswith(fam + " "):
        b = b[len(fam) + 1:]
    return b


# --------------------------------------------------------------------------
# WCVP through GBIF
# --------------------------------------------------------------------------
def wcvp_search(genus):
    out, offset = [], 0
    while True:
        d = get_json(GBIF + "/species/search",
                     {"datasetKey": WCVP_DATASET, "q": genus, "limit": 500, "offset": offset})
        rs = d.get("results", [])
        for r in rs:
            cn = r.get("canonicalName") or ""
            if cn == genus or cn.startswith(genus + " "):
                out.append(r)
        if d.get("endOfRecords", True) or not rs:
            break
        offset += len(rs)
    return out


def wcvp_full(key):
    return get_json(GBIF + "/species/%d" % key)


def wcvp_synonyms(key):
    out, offset = [], 0
    while True:
        d = get_json(GBIF + "/species/%d/synonyms" % key, {"limit": 200, "offset": offset})
        rs = d.get("results", [])
        out.extend(rs)
        if d.get("endOfRecords", True) or not rs:
            break
        offset += len(rs)
    return out


def wcvp_distribution(key):
    out, offset = [], 0
    while True:
        d = get_json(GBIF + "/species/%d/distributions" % key, {"limit": 500, "offset": offset})
        rs = d.get("results", [])
        out.extend(rs)
        if d.get("endOfRecords", True) or not rs:
            break
        offset += len(rs)
    return out


def wcvp_dataset_meta():
    d = get_json(GBIF + "/dataset/" + WCVP_DATASET)
    return {"title": d.get("title"), "pubDate": (d.get("pubDate") or "")[:10],
            "doi": d.get("doi") or WCVP_DOI}


STATUS_LABEL = {"ACCEPTED": "Accepted", "SYNONYM": "Synonym",
                "HOMOTYPIC_SYNONYM": "Synonym", "HETEROTYPIC_SYNONYM": "Synonym",
                "PROPARTE_SYNONYM": "Synonym", "DOUBTFUL": "Unplaced",
                "MISAPPLIED": "Misapplied"}


RANK_MARKER = {"SUBSPECIES": "subsp.", "VARIETY": "var.", "SUBVARIETY": "subvar.",
               "FORM": "f.", "SUBFORM": "subf."}


def display_name(canonical, rank):
    """GBIF's canonicalName drops the rank marker ('X y z'); put it back."""
    parts = canonical.split()
    marker = RANK_MARKER.get((rank or "").upper())
    if marker and len(parts) == 3:
        return "%s %s %s %s" % (parts[0], parts[1], marker, parts[2])
    return canonical


def authority_differs(ipni_auth, powo_auth):
    """IPNI's `authors` often omits the basionym bracket that POWO shows -
    '(Schott) Engl.' vs 'Engl.' - which is not a disagreement."""
    a = ipni_auth.replace(" ", "")
    b = powo_auth.replace(" ", "")
    if not a or not b or a == b:
        return False
    if ")" in b and not a.startswith("("):
        b = b.split(")", 1)[1]
    return a != b


def year_of(s):
    if not s:
        return None
    m = YEAR_PAREN.search(s)
    if m:
        return int(m.group(1))
    m = YEAR_ANY.search(s)
    return int(m.group(1)) if m else None


# --------------------------------------------------------------------------
# TDWG code -> POWO display name, unioned over every genus in genus-geo.json
# --------------------------------------------------------------------------
def tdwg_names():
    p = os.path.join(REPO, "docs", "genus-geo.json")
    table = {}
    if os.path.exists(p):
        with io.open(p, encoding="utf-8") as f:
            g = json.load(f)
        for entry in (g.get("genera") or {}).values():
            for name, code in (entry.get("codes") or {}).items():
                table.setdefault(code, name)
    return table


# --------------------------------------------------------------------------
# Reconcile
# --------------------------------------------------------------------------
def build(genus):
    print("IPNI  : searching genus:%s" % genus)
    ipni = ipni_names(genus)
    print("        %d name records" % len(ipni))

    print("WCVP  : searching %s in the GBIF copy" % genus)
    wcvp = {r["key"]: r for r in wcvp_search(genus)}
    accepted_keys = [k for k, r in wcvp.items()
                     if r.get("taxonomicStatus") == "ACCEPTED" and r.get("rank") == "SPECIES"]
    print("        %d records, %d accepted species" % (len(wcvp), len(accepted_keys)))

    # Synonyms of each accepted species, including names in OTHER genera.
    for k in accepted_keys:
        for s in wcvp_synonyms(k):
            wcvp.setdefault(s["key"], s)

    # Full records: POWO URL (=> IPNI id), range summary, basionym.
    full = {}
    for k in list(wcvp):
        full[k] = wcvp_full(k)

    print("WCVP  : %d records after pulling each accepted species' synonymy" % len(wcvp))

    codes = tdwg_names()
    ipni_by_id = {r["id"]: r for r in ipni}
    # Names from OTHER genera that POWO sinks into this genus are outside the
    # IPNI genus search; fetch their IPNI records by id so they carry the same
    # protologue and basionym detail.
    for k, f in full.items():
        m = IPNI_LSID.search(f.get("references") or "")
        if m and m.group(1) not in ipni_by_id:
            try:
                ipni_by_id[m.group(1)] = get_json("https://www.ipni.org/api/1/n/" + m.group(1))
            except Exception as e:
                print("  IPNI record %s not fetched: %s" % (m.group(1), str(e)[:60]))
    used_ipni = set()
    names = []

    def wcvp_entry(k):
        r, f = wcvp[k], full.get(k, {})
        m = IPNI_LSID.search(f.get("references") or "")
        ipni_id = m.group(1) if m else None
        ir = ipni_by_id.get(ipni_id) if ipni_id else None
        if ir:
            used_ipni.add(ipni_id)
        status = STATUS_LABEL.get(r.get("taxonomicStatus") or "", r.get("taxonomicStatus") or "")
        canonical = display_name(r.get("canonicalName") or "", r.get("rank") or "")
        notes = []
        if ir and ir.get("name") and ir["name"] != canonical:
            notes.append("IPNI spells it '%s'" % ir["name"])
        if ir and authority_differs(ir.get("authors") or "", r.get("authorship") or ""):
            notes.append("IPNI authority '%s'" % ir["authors"])
        year = None
        if ir and ir.get("publicationYear"):
            year = int(ir["publicationYear"])
        if year is None:
            year = year_of(r.get("publishedIn")) or (year_of(ir.get("reference")) if ir else None)
        protologue = ipni_protologue(ir) if ir else ""
        if not protologue:
            protologue = (r.get("publishedIn") or "").strip()
        basionym = (f.get("basionym") or "").strip() or (ipni_basionym(ir) if ir else "")
        return {
            "name": canonical,
            "authority": (r.get("authorship") or (ir.get("authors") if ir else "") or "").strip(),
            "rank": (r.get("rank") or "").lower(),
            "year": year,
            "protologue": protologue,
            "basionym": basionym,
            "originalRemarks": (ir.get("originalRemarks") or "").strip() if ir else "",
            "ipniId": ipni_id,
            "ipniUrl": "https://www.ipni.org/n/" + ipni_id if ipni_id else "",
            "powoUrl": (f.get("references") or "").strip(),
            "wcvpId": r.get("taxonID") or f.get("taxonID") or "",
            "wfoId": (ir.get("wfoId") or "") if ir else "",
            "gbifKey": k,
            "bhlLink": (ir.get("bhlLink") or "") if ir else "",
            "powoStatus": status,
            "acceptedName": (r.get("accepted") or "").strip() if status != "Accepted" else "",
            "acceptedKey": r.get("acceptedKey"),
            "rangeSummary": (f.get("remarks") or "").strip() if status == "Accepted" else "",
            "inGenus": canonical == genus or canonical.startswith(genus + " "),
            "hybrid": bool((ir or {}).get("hybrid")) or HYBRID in canonical,
            "sources": ["IPNI", "POWO/WCVP"] if ir else ["POWO/WCVP"],
            "notes": notes,
        }

    for k in wcvp:
        names.append(wcvp_entry(k))

    # IPNI-only names: nothing in WCVP points at this IPNI id.
    for r in ipni:
        if r["id"] in used_ipni:
            continue
        names.append({
            "name": (r.get("name") or "").strip(),
            "authority": (r.get("authors") or "").strip(),
            "rank": {"spec.": "species", "gen.": "genus", "var.": "variety",
                     "subsp.": "subspecies", "f.": "form"}.get(r.get("rank"), r.get("rank") or ""),
            "year": int(r["publicationYear"]) if r.get("publicationYear") else year_of(r.get("reference")),
            "protologue": ipni_protologue(r),
            "basionym": ipni_basionym(r),
            "originalRemarks": (r.get("originalRemarks") or "").strip(),
            "ipniId": r["id"],
            "ipniUrl": "https://www.ipni.org/n/" + r["id"],
            "powoUrl": "https://powo.science.kew.org/taxon/urn:lsid:ipni.org:names:" + r["id"],
            "wcvpId": "",
            "wfoId": (r.get("wfoId") or ""),
            "gbifKey": None,
            "bhlLink": (r.get("bhlLink") or ""),
            "powoStatus": "Not in POWO" if not r.get("inPowo") else "Not in the WCVP copy",
            "acceptedName": "",
            "acceptedKey": None,
            "rangeSummary": "",
            "inGenus": True,
            "hybrid": bool(r.get("hybrid")) or HYBRID in (r.get("name") or ""),
            "sources": ["IPNI"],
            "notes": [],
        })

    # Distribution for accepted species.
    dist = {}
    for k in accepted_keys:
        native, introduced, other = [], [], []
        for d in wcvp_distribution(k):
            code = (d.get("locationId") or "").replace("TDWG:", "")
            place = codes.get(code) or d.get("locality") or code
            # The WCVP copy on GBIF leaves establishmentMeans EMPTY for native
            # units and says INTRODUCED for the rest (measured 2026-09-05).
            em = (d.get("establishmentMeans") or "NATIVE").upper()
            (native if em == "NATIVE" else introduced if em == "INTRODUCED" else other).append(place)
        dist[k] = {"native": sorted(set(native)), "introduced": sorted(set(introduced)),
                   "other": sorted(set(other))}

    by_key = {n["gbifKey"]: n for n in names if n["gbifKey"]}
    for n in names:
        if n["powoStatus"] == "Accepted" and n["gbifKey"] in dist:
            n.update(dist[n["gbifKey"]])
            n["synonyms"] = sorted(
                "%s %s" % (m["name"], m["authority"]) for m in names
                if m["acceptedKey"] == n["gbifKey"])
        else:
            n["native"], n["introduced"], n["other"], n["synonyms"] = [], [], [], []

    # Species that a synonym points to but which is not itself accepted in
    # this genus (i.e. sunk into another genus) keep the accepted string only.
    def sort_key(n):
        return (0 if n["rank"] == "genus" else 1, n["name"].lower(), n["authority"])
    names.sort(key=sort_key)

    accepted = [n for n in names if n["powoStatus"] == "Accepted" and n["rank"] == "species"
                and n["inGenus"] and not n["hybrid"]]
    hybrids = [n for n in names if n["hybrid"]]
    synonyms = [n for n in names if n not in accepted and n not in hybrids and n["rank"] != "genus"]
    synonyms.sort(key=lambda n: (0 if n["inGenus"] else 1, n["name"].lower(), n["authority"]))

    meta = wcvp_dataset_meta()
    discrepancies = []
    for n in names:
        for t in n["notes"]:
            discrepancies.append("%s %s: %s" % (n["name"], n["authority"], t))
        if n["sources"] == ["IPNI"]:
            discrepancies.append("%s %s (IPNI %s): in IPNI only - %s" % (
                n["name"], n["authority"], n["ipniId"], n["powoStatus"]))
        if n["sources"] == ["POWO/WCVP"] and n["inGenus"]:
            discrepancies.append("%s %s: in POWO/WCVP only, IPNI has no record with this id" % (
                n["name"], n["authority"]))

    return {
        "genus": genus,
        "generated": datetime.datetime.now(datetime.timezone.utc).strftime("%Y-%m-%d"),
        "sources": {
            "ipni": IPNI_SEARCH + "?q=genus:" + genus,
            "wcvpViaGbif": GBIF + "/species/search?datasetKey=" + WCVP_DATASET + "&q=" + genus,
            "wcvpDataset": meta,
            "powoNote": "powo.science.kew.org/api answers 403 (Cloudflare challenge) to scripts; "
                        "the WCVP copy on GBIF is the same checklist POWO serves. Each row carries "
                        "its POWO URL for checking on the site.",
        },
        "counts": {
            "ipniRecords": len(ipni),
            "wcvpRecords": len(wcvp),
            "acceptedSpecies": len(accepted),
            "hybrids": len(hybrids),
            "otherNames": len(synonyms),
            "ipniOnly": sum(1 for n in names if n["sources"] == ["IPNI"]),
        },
        "discrepancies": discrepancies,
        "accepted": accepted,
        "hybrids": hybrids,
        "synonyms": synonyms,
        "names": names,
    }


# --------------------------------------------------------------------------
# Cross-check against what the repo already carries
# --------------------------------------------------------------------------
def crosscheck(doc):
    slug = doc["genus"].lower()
    out = []
    p = os.path.join(REPO, "data", "species-years.json")
    if os.path.exists(p):
        with io.open(p, encoding="utf-8") as f:
            sy = json.load(f)
        g = (sy.get("genera") or {}).get(slug)
        if g:
            have = {n: int(y) for y, ns in g["years"].items() for n in ns}
            mine = {n["name"]: n["year"] for n in doc["accepted"]}
            only_repo = sorted(set(have) - set(mine))
            only_here = sorted(set(mine) - set(have))
            diff_year = sorted(n for n in set(have) & set(mine) if have[n] != mine[n])
            out.append("data/species-years.json (POWO export %s): %d species there, %d here"
                       % (sy.get("exportsNewest", "?"), len(have), len(mine)))
            out.append("  only in the export : %s" % (", ".join(only_repo) or "none"))
            out.append("  only in this pull  : %s" % (", ".join(only_here) or "none"))
            out.append("  year differs       : %s" % (
                ", ".join("%s (%s vs %s)" % (n, have[n], mine[n]) for n in diff_year) or "none"))
        else:
            out.append("data/species-years.json: genus not present")
    p = os.path.join(REPO, "docs", "genus-geo.json")
    if os.path.exists(p):
        with io.open(p, encoding="utf-8") as f:
            gg = json.load(f)
        e = (gg.get("genera") or {}).get(doc["genus"])
        if e:
            sp = e.get("speciesPlaces") or {}
            out.append("docs/genus-geo.json: speciesTotalPowo %s, %d species with places"
                       % (e.get("speciesTotalPowo"), len(sp)))
            for n in doc["accepted"]:
                ep = n["name"].split(" ", 1)[1] if " " in n["name"] else n["name"]
                if ep in sp and sorted(sp[ep]) != sorted(n["native"]):
                    out.append("  %s native range differs: geo %s | pull %s"
                               % (n["name"], sorted(sp[ep]), n["native"]))
    return out


# --------------------------------------------------------------------------
# Output
# --------------------------------------------------------------------------
SPECIES_COLS = [
    ("SPECIES NAME", "name"), ("AUTHORITY", "authority"), ("YEAR DESCRIBED", "year"),
    ("PROTOLOGUE", "protologue"), ("BASIONYM", "basionym"),
    ("TYPE / ORIGINAL REMARKS (IPNI)", "originalRemarks"),
    ("POWO RANGE SUMMARY", "rangeSummary"),
    ("NATIVE RANGE (TDWG L3)", "native"), ("INTRODUCED RANGE (TDWG L3)", "introduced"),
    ("SYNONYMS (POWO)", "synonyms"),
    ("IPNI ID", "ipniId"), ("IPNI URL", "ipniUrl"), ("POWO URL", "powoUrl"),
    ("WCVP ID", "wcvpId"), ("WFO ID", "wfoId"), ("BHL (PROTOLOGUE SCAN)", "bhlLink"),
    ("SOURCES", "sources"),
    # Editorial columns - blank on purpose; the sources never fill these.
    ("AROIDPEDIA STATUS", None), ("AROIDPEDIA RULING / NOTES", None), ("PAGE URL", None),
]
HYBRID_COLS = [
    ("HYBRID NAME", "name"), ("AUTHORITY", "authority"), ("YEAR DESCRIBED", "year"),
    ("PROTOLOGUE", "protologue"), ("POWO STATUS", "powoStatus"), ("ACCEPTED NAME (POWO)", "acceptedName"),
    ("IPNI ID", "ipniId"), ("IPNI URL", "ipniUrl"), ("POWO URL", "powoUrl"), ("SOURCES", "sources"),
    ("PARENTAGE", None), ("AROIDPEDIA NOTES", None),
]
SYNONYM_COLS = [
    ("NAME", "name"), ("AUTHORITY", "authority"), ("RANK", "rank"), ("YEAR", "year"),
    ("PROTOLOGUE", "protologue"), ("POWO STATUS", "powoStatus"),
    ("ACCEPTED NAME (POWO)", "acceptedName"), ("RELATION", "relation"),
    ("BASIONYM", "basionym"), ("TYPE / ORIGINAL REMARKS (IPNI)", "originalRemarks"),
    ("IPNI ID", "ipniId"), ("IPNI URL", "ipniUrl"), ("POWO URL", "powoUrl"),
    ("WCVP ID", "wcvpId"), ("SOURCES", "sources"), ("NOTES", "notes"),
]
ALL_COLS = [
    ("NAME", "name"), ("AUTHORITY", "authority"), ("RANK", "rank"), ("YEAR", "year"),
    ("PROTOLOGUE", "protologue"), ("POWO STATUS", "powoStatus"),
    ("ACCEPTED NAME (POWO)", "acceptedName"), ("IN THIS GENUS", "inGenus"), ("HYBRID", "hybrid"),
    ("IPNI ID", "ipniId"), ("WCVP ID", "wcvpId"), ("WFO ID", "wfoId"),
    ("IPNI URL", "ipniUrl"), ("POWO URL", "powoUrl"), ("SOURCES", "sources"), ("NOTES", "notes"),
]


def relation(n, genus):
    if n["powoStatus"] == "Accepted":
        return "accepted in another rank/genus" if not n["inGenus"] else "accepted"
    if n["powoStatus"] == "Unplaced":
        return "%s name POWO leaves unplaced" % genus
    if n["sources"] == ["IPNI"]:
        return "IPNI record only"
    acc = n["acceptedName"] or ""
    if n["inGenus"] and acc.startswith(genus + " "):
        return "%s name sunk into an accepted %s species" % (genus, genus)
    if n["inGenus"]:
        return "%s name now accepted in another genus" % genus
    return "name from another genus sunk into an accepted %s species" % genus


def cell(v):
    if v is None:
        return ""
    if isinstance(v, bool):
        return "yes" if v else "no"
    if isinstance(v, list):
        return "; ".join(str(x) for x in v)
    return v


def rows_for(items, cols, genus):
    out = []
    for n in items:
        n = dict(n, relation=relation(n, genus))
        out.append([cell(n.get(k)) if k else "" for _, k in cols])
    return out


def write_sheet(wb, title, cols, rows, widths=None, first=False):
    ws = wb.active if first else wb.create_sheet()
    ws.title = title
    ws.append([h for h, _ in cols])
    for r in rows:
        ws.append(r)
    head = Font(bold=True)
    fill = PatternFill("solid", fgColor="E8EFE6")
    for c in ws[1]:
        c.font, c.fill = head, fill
        c.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for i, (h, _) in enumerate(cols, 1):
        w = (widths or {}).get(h)
        if w is None:
            longest = max([len(str(h))] + [len(str(r[i - 1])) for r in rows] or [len(h)])
            w = min(max(12, longest + 2), 60)
        ws.column_dimensions[get_column_letter(i)].width = w
    return ws


def write_outputs(doc, checks):
    genus = doc["genus"]
    os.makedirs(OUT_DIR, exist_ok=True)
    wb = Workbook()

    sp_rows = rows_for(doc["accepted"], SPECIES_COLS, genus)
    write_sheet(wb, "SPECIES", SPECIES_COLS, sp_rows, first=True,
                widths={"SPECIES NAME": 34, "AUTHORITY": 36, "YEAR DESCRIBED": 10,
                        "PROTOLOGUE": 46, "SYNONYMS (POWO)": 60, "NATIVE RANGE (TDWG L3)": 50,
                        "AROIDPEDIA STATUS": 20, "AROIDPEDIA RULING / NOTES": 40, "PAGE URL": 30})
    write_sheet(wb, "HYBRIDS", HYBRID_COLS, rows_for(doc["hybrids"], HYBRID_COLS, genus),
                widths={"HYBRID NAME": 34, "AUTHORITY": 30, "PROTOLOGUE": 46, "PARENTAGE": 40,
                        "AROIDPEDIA NOTES": 40})
    syn_rows = rows_for(doc["synonyms"], SYNONYM_COLS, genus)
    write_sheet(wb, "SYNONYMS", SYNONYM_COLS, syn_rows,
                widths={"NAME": 40, "AUTHORITY": 36, "PROTOLOGUE": 52, "RELATION": 56})
    write_sheet(wb, "ALL NAMES", ALL_COLS, rows_for(doc["names"], ALL_COLS, genus),
                widths={"NAME": 40, "AUTHORITY": 36, "PROTOLOGUE": 52})

    ws = wb.create_sheet("SOURCES")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 120
    c = doc["counts"]
    m = doc["sources"]["wcvpDataset"]
    lines = [
        ("Genus", genus),
        ("Generated", doc["generated"] + " (UTC), by scripts/build-species-base.py"),
        ("Regenerate", "python scripts/build-species-base.py %s   (add --refresh to ignore the cache)" % genus),
        ("", ""),
        ("IPNI", doc["sources"]["ipni"]),
        ("POWO / WCVP", "%s - %s release, DOI %s, read through GBIF: %s" % (
            m.get("title"), m.get("pubDate"), m.get("doi"), doc["sources"]["wcvpViaGbif"])),
        ("POWO note", doc["sources"]["powoNote"]),
        ("Distribution", "POWO/WCVP native and introduced TDWG level-3 units per accepted species "
                         "(GBIF /species/{key}/distributions), shown with POWO's place names where "
                         "docs/genus-geo.json knows the code."),
        ("", ""),
        ("IPNI name records", c["ipniRecords"]),
        ("POWO/WCVP records", "%d (genus search plus each accepted species' full synonymy)" % c["wcvpRecords"]),
        ("Accepted species (POWO)", c["acceptedSpecies"]),
        ("Hybrids", c["hybrids"]),
        ("Other names (SYNONYMS sheet)", c["otherNames"]),
        ("IPNI-only names", c["ipniOnly"]),
        ("", ""),
        ("House rule", "POWO is the base list, not the verdict: a published journal article trumps it. "
                       "This workbook is a faithful pull of the two sources; editorial rulings go in the "
                       "SPECIES sheet's AROIDPEDIA columns and in data/taxon-rulings.json."),
        ("SPECIES sheet", "Accepted species only, hybrids on HYBRIDS - the same shape as the POWO export "
                          "workbooks, so scripts/extract-species-years.py can read this file unchanged "
                          "(it looks for the SPECIES NAME and YEAR DESCRIBED headers)."),
        ("", ""),
    ]
    lines.append(("Cross-check vs repo", checks[0] if checks else "nothing to compare"))
    for l in checks[1:]:
        lines.append(("", l))
    lines.append(("", ""))
    lines.append(("Discrepancies (%d)" % len(doc["discrepancies"]),
                  doc["discrepancies"][0] if doc["discrepancies"] else "none"))
    for d in doc["discrepancies"][1:]:
        lines.append(("", d))
    for a, b in lines:
        ws.append([a, b])
    for row in ws.iter_rows(min_row=1, max_col=1):
        row[0].font = Font(bold=True)

    xlsx = os.path.join(OUT_DIR, genus + ".xlsx")
    wb.save(xlsx)

    def write_csv(path, cols, rows):
        with io.open(path, "w", encoding="utf-8", newline="") as f:
            w = csv.writer(f)
            w.writerow([h for h, _ in cols])
            w.writerows(rows)

    write_csv(os.path.join(OUT_DIR, genus + "-species.csv"), SPECIES_COLS, sp_rows)
    write_csv(os.path.join(OUT_DIR, genus + "-synonyms.csv"), SYNONYM_COLS, syn_rows)

    slim = dict(doc)
    slim["crosscheck"] = checks
    with io.open(os.path.join(OUT_DIR, genus + "-names.json"), "w", encoding="utf-8", newline="\n") as f:
        f.write(json.dumps(slim, indent=1, ensure_ascii=False) + "\n")
    return xlsx


def main():
    global _refresh
    ap = argparse.ArgumentParser(description="Build a genus species base from POWO/WCVP + IPNI.")
    ap.add_argument("genus", help="Genus name, capitalised, e.g. Cyrtosperma")
    ap.add_argument("--refresh", action="store_true", help="ignore the on-disk API cache")
    a = ap.parse_args()
    _refresh = a.refresh
    genus = a.genus.strip().capitalize()

    doc = build(genus)
    checks = crosscheck(doc)
    xlsx = write_outputs(doc, checks)

    c = doc["counts"]
    print()
    print("accepted species  : %d" % c["acceptedSpecies"])
    for n in doc["accepted"]:
        print("  %-4s %s %s" % (n["year"] or "----", n["name"], n["authority"]))
    print("hybrids           : %d" % c["hybrids"])
    print("other names       : %d   (IPNI-only %d)" % (c["otherNames"], c["ipniOnly"]))
    print("API calls made    : %d   (cache: %s)" % (_calls, CACHE))
    if checks:
        print("\ncross-check:")
        for l in checks:
            print("  " + l)
    if doc["discrepancies"]:
        print("\ndiscrepancies between the two sources (%d):" % len(doc["discrepancies"]))
        for d in doc["discrepancies"]:
            print("  . " + d)
    print("\n->", xlsx)


if __name__ == "__main__":
    main()
